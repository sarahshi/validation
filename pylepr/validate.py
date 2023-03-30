import logging

import numpy as np
import pandas as pd
from openpyxl.utils.cell import get_column_letter

log_filename = 'validation.log'
logging.basicConfig(filename=log_filename,
                    filemode='w',
                    format="---> %(levelname)s..(%(funcName)s):: %(message)s", 
                    force = True)

def print_log_file(log_filename=log_filename):
    with open(log_filename, 'r') as fin:
        print(fin.read())
        
        


CHEM_DATA_INFO = {
    'sheetname': '6 Run Products',
    'header_row_num': 4,
    'metadata_header_row_num': 2,
    'chem_dat_col_index': 13
}

def extract_chem_dat(upload_data, format=CHEM_DATA_INFO):
    run_products = upload_data[format['sheetname']]
    
    
    run_names = run_products.iloc[format['header_row_num']+1:,0]

    dat = run_products.iloc[:,format['chem_dat_col_index']:]
    dat.columns = dat.iloc[0]
    dat = dat.iloc[1:]
    chem_dat_info = dat.iloc[:2]
    chem_dat_info.index = ['method_id','unit']

    chem_dat = dat.iloc[format['header_row_num']:]
    chem_dat
    chem_dat.index = run_names

    return chem_dat, chem_dat_info


# def _get_chem_info_cell(cell_data, format):
#     col_num = (format['chem_dat_col_index'] + 
#                cell_data['col_ind']+1)
#     col_str = get_column_letter(col_num)
#     row_num = (format['header_row_num'] +
#                format['metadata_header_row_num'] +
#                cell_data['row_ind']+1)
#     return f'{col_str}{row_num}'

def _validate_chem_error_columns(chem_dat_info, format):
    columns = chem_dat_info.columns
    meas_cols = [col for col in columns if not col.endswith('_err') ]
    for col in meas_cols:
        # col_str = get_column_letter(col)
        if col+'_err' not in columns:
            logging.error(f"'{col}_err' missing from chemistry data columns")
            

def _validate_chem_units(chem_dat_info, format):
    for col_ind, (col, dat) in enumerate(chem_dat_info.T.iterrows()):
        col_num = format['chem_dat_col_index']+col_ind+1
        col_str = get_column_letter(col_num)
        row_num = 4
        cell_id = f"{col_str}{row_num}"
        if dat.unit is np.nan:
            logging.critical(f"<<cell {cell_id}>>: '{col}' does not provide any units")
            
def _validate_chem_method(chem_dat_info, format):
    for col_ind, (col, dat) in enumerate(chem_dat_info.T.iterrows()):
        col_num = format['chem_dat_col_index']+col_ind+1
        col_str = get_column_letter(col_num)
        row_num = 3
        cell_id = f"{col_str}{row_num}"
        if dat.method_id is np.nan:
            logging.critical(f"<<cell {cell_id}>>: '{col}' does not provide any method id")
            

def validate_chem_data_info(chem_dat_info,
                            format=CHEM_DATA_INFO):
    _validate_chem_error_columns(chem_dat_info, format)
    _validate_chem_units(chem_dat_info, format)
    _validate_chem_method(chem_dat_info, format)
    
    

        
 
def _chem_not_detected_not_valid(cell_data, format):
    val, chem, run_id = [cell_data[key] for key in 
                         ['val','chem','run_id']]
        
    cell_id = _get_chem_data_cell(cell_data, format)
    
    
    if val=='nd':
        logging.error(f"<<cell {cell_id}>>: '{val}', the '{chem}' value for exp_run '{run_id}' is not valid. If not detected use vocabulary 'bdl'")
        return True
    
    return False

def _chem_not_measured_not_valid(cell_data, format):
    val, chem, run_id = [cell_data[key] for key in 
                         ['val','chem','run_id']]
        
    cell_id = _get_chem_data_cell(cell_data, format)
    
    
    if val=='-':
        logging.error(f"<<cell {cell_id}>>: '{val}', the '{chem}' value for exp_run '{run_id}' is not valid. If not measured leave entry blank")
        return True
    
    return False

def _chem_measurement_limit_not_valid(cell_data, format):
    val, chem, run_id = [cell_data[key] for key in 
                         ['val','chem','run_id']]
    
    cell_id = _get_chem_data_cell(cell_data, format)
    
    if type(val) is not str:
        return False
    
    if val.startswith('>') or val.startswith('<'):
        logging.error(f"<<cell {cell_id}>>: '{val}', the '{chem}' value for exp_run '{run_id}' is not valid. Instead give just the value and indicate limit using field '????, Ask roger'")
        return True
    
    return False

def _get_chem_data_cell(cell_data, format):
    col_num = (format['chem_dat_col_index'] + 
               cell_data['col_ind']+1)
    col_str = get_column_letter(col_num)
    row_num = (format['header_row_num'] +
               format['metadata_header_row_num'] +
               cell_data['row_ind']+1)
    return f'{col_str}{row_num}'
        
def _numeric_chem_data_not_valid(cell_data, format):

    val, chem, run_id = [cell_data[key] for key in 
                         ['val','chem','run_id']]

    cell_id = _get_chem_data_cell(cell_data, format)
    if type(val) is str:
        logging.error(f"<<cell {cell_id}>>: '{chem}' value for exp_run '{run_id}' is invalid. '{val}' is not a valid number.")
        return True
    
    return False


def validate_chem_data(chem_dat, format=CHEM_DATA_INFO):
    for col_ind, (ichem_col, ichem_dat) in enumerate(chem_dat.T.iterrows()):
        chem = ichem_dat.name
        for row_ind, (run_id, val) in enumerate(ichem_dat.items()):
            cell_data = {'val':val, 'chem':chem, 
                         'run_id':run_id, 'col_ind':col_ind,
                         'row_ind':row_ind}
            if _chem_not_detected_not_valid(cell_data, format):
                continue
            
            if _chem_not_measured_not_valid(cell_data, format):
                continue
            
            if _chem_measurement_limit_not_valid(cell_data, format):
                continue
            
            
            _numeric_chem_data_not_valid(cell_data, format)
            
            
def validate_upload(upload_data):
    chem_dat, chem_dat_info = extract_chem_dat(upload_data)
    validate_chem_data_info(chem_dat_info)
    validate_chem_data(chem_dat)
    print_log_file()